"""Contains the AlaskaProject class.

Alaska is organized into "projects." These project hold all information about
a specific experiment. The ultimate goal of a project is to perform
differential expression analysis and identify differentially expressed genes
among samples.
"""

__author__ = 'Kyung Hoi (Joseph) Min'
__copyright__ = 'Copyright 2017 WormLabCaltech'
__credits__ = ['David Angeles', 'Raymond Lee', 'Juancarlos Chan']
__license__ = "MIT"
__version__ = "alpha"
__maintainer__ = "Kyung Hoi (Joseph) Min"
__email__ = "kmin@caltech.edu"
__status__ = "alpha"

import os
import json
import pandas as pd
import warnings as w
import datetime as dt
from copy import copy
import openpyxl as opx
from collections import defaultdict
from pyunpack import Archive
from BashWriter import BashWriter
from AlaskaSample import AlaskaSample
# from multiprocessing import Process
from Alaska import Alaska

class AlaskaProject(Alaska):
    """
    AlaskaProject. Class to wrap all project data.
    """

    def __init__(self, _id):
        """
        AlaskaProject constructor. Must receive id.
        """
        self.id = _id
        self.dir = '{}/{}'.format(Alaska.PROJECTS_DIR, _id)
        self.qc_dir = '{}/{}'.format(self.dir, Alaska.QC_DIR)
        self.raw_dir = '{}/{}'.format(self.dir, Alaska.RAW_DIR)
        self.align_dir = '{}/{}'.format(self.dir, Alaska.ALIGN_DIR)
        self.diff_dir = '{}/{}'.format(self.dir, Alaska.DIFF_DIR)
        self.temp_dir = '{}/{}'.format(self.dir, Alaska.TEMP_DIR)
        self.jobs = [] # jobs related to this project
        self.raw_reads = {}
        self.samples = {}
        self.design = 1 # 1: single-factor, 2: two-factor
        self.controls = [] # controls
        self.factors = []

        self.progress = 0 # int to denote current analysis progress

        self.meta = {} # variable for all metadata
        # from GEO submission template
        self.meta['title'] = ''
        self.meta['abstract'] = ''
        self.meta['corresponding'] = {
            'email': '',
            'name': ''
        }
        self.meta['contributors'] = []
        self.meta['SRA_center_code'] = ''
        # end from GEO submission template
        self.meta['datetime'] = dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    def fetch_reads(self):
        """
        Simply fetches all files & folders in the raw reads directory in JSON format.
        """
        reads = []
        # walk through the raw directory
        for root, dirs, files in os.walk(self.raw_dir):
            for fname in files:
                # skip files that are not one of the recognized extensions.
                extensions = Alaska.RAW_EXT + Alaska.ARCH_EXT
                if not fname.endswith(extensions):
                    continue

                # otherwise, save info about the file
                path = '{}/{}/{}'.format(Alaska.ROOT_PATH, root, fname)
                folder = root.replace(self.raw_dir, '')
                filename = fname
                size = os.path.getsize(path) / (1024 ** 2)
                read = {
                    'folder': folder,
                    'filename': fname,
                    'size': '{} MB'.format(round(size,1)),
                    'path': path
                }
                reads.append(read)
        return reads

    def get_raw_reads(self):
        """
        Retrieves list of uploaded sample files. Unpacks archive if detected.
        """
        unpack = []
        for root, dirs, files in os.walk(self.raw_dir):
            for fname in files:
                # if the file name does not end in a known raw read extension,
                # and is a known (and unpackable) archive, add to list
                if not fname.endswith(Alaska.RAW_EXT) and fname.endswith(Alaska.ARCH_EXT):
                    unpack.append('{}/{}'.format(root, fname))

        # if files need to be unpack_reads
        if not len(unpack) == 0:
            self.out('{}: unpacking required'.format(self.id))
            for fname in unpack:
                self.out('{}: unpacking {}'.format(self.id, fname))
                try:
                    # p = Process(target=self.unpack_reads, args=(fname,))
                    # p.daemon = True
                    # p.start()
                    # p.join()
                    self.unpack_reads(fname)
                except Exception as e:
                    print(str(e))
                    raise Exception('{}: exception occured while unpacking {}'.format(self.id, fname))
            self.out('{}: unpacking finished'.format(self.id))

        # walk through raw reads directory
        self.raw_reads = {}
        for root, dirs, files in os.walk(self.raw_dir):
            # go straight to deepest directory
            if not len(dirs) == 0:
                continue

            reads = {} # list to contain read files for each directory
            for fname in files:
                # only files ending with certain extensions
                # and not directly located in raw read directory should be added
                if fname.endswith(Alaska.RAW_EXT) and '{}/{}'.format(root, fname) not in unpack:
                    full_path = '{}/{}'.format(root, fname)

                    # remove project folder from root
                    split = root.split('/')
                    split.remove(Alaska.PROJECTS_DIR)
                    split.remove(self.id)

                    path = '{}/{}'.format('/'.join(split), fname)
                    size = os.path.getsize(full_path)
                    md5 = None

                    read = {}
                    read['size'] = round(size / (1024 ** 2),1)
                    read['md5'] = md5

                    reads[path] = read


            # assign list to dictionary
            if not len(reads) == 0:
                self.raw_reads[root.replace(self.raw_dir, '')[1:]] = reads
            else:
                raise Exception('{}: raw reads folder is empty!'.format(self.id))

    def unpack_reads(self, fname):
        """
        Unpacks read archive.
        """
        Archive(fname).extractall(fname + '_extracted', auto_create_dir=True)

    def infer_samples(self, f, temp=None, md5=True):
        """
        Infers samples from raw reads.
        Assumes each sample is in separate folders.
        Only to be called when raw reads is not empty.
        """
        # TODO: add way to infer single- or pair-end read
        w.warn('{}: Alaska is currently unable to infer paired-end samples'
                .format(self.id), Warning)

        # # make sure that md5 checksums have been calculated
        # if md5 and len(self.chk_md5) == 0:
        #     raise Exception('{}: MD5 checksums have not been calculated'.format(self.id))

        # loop through each folder with sample
        self.samples = {}
        for folder, reads in self.raw_reads.items():
            _id = 'AS{}'.format(f())
            sample = AlaskaSample(_id, folder)

            if temp is not None: # if temporary variable is given
                temp[_id] = sample

            self.out('{}: new sample created with id {}'.format(self.id, _id))

            for read, item in reads.items():
                sample.reads[read] = item

            sample.projects.append(self.id)
            self.samples[_id] = sample

    def analyze_reads(self):
        """
        Analyzes reads to infer whether samples are single or paired end.
        """
        # TODO: implement

    def check(self):
        """
        Checks all data.
        """
        # Have to check: design vs sample
        w.warn('{}: Alaska is currently unable to verify experimental designs'.format(self.id),
                Warning)

        # ctrl_chars = {}
        # # first, get control characteristics
        # for _id, char in self.ctrls.items():
        #     if char not in ctrl_chars:
        #         ctrl_chars[char] = self.samples[_id].meta[char]
        #
        # # check controls & their characteristics
        # for _id, sample in self.samples.items():
        #     if _id in self.ctrls:
        #         char = self.ctrls[_id]
        #         if not sample.meta[char] == ctrl_chars[char]:
        #             msg = '{}: control sample {} does not have {} : {}' \
        #                     .format(self.id, _id, char, ctrl_chars[char])
        #             raise Exception(msg)
        #     else:
        #         if sample.meta[char] == ctrl_chars[char]:
        #             msg = '{}: non-control sample {} has {} : {}' \
        #                     .format(self.id, _id, char, ctrl_chars[char])
        #             raise Exception(msg)



        # # first check controls have matching control factors
        # # then check if non-controls have different factors
        # # TODO: check other data and ensure other factors are the same
        # for _id, sample in self.samples.items():
        #     # check controls have matching control factors
        #     if _id in self.ctrls:
        #         for __id, char in self.ctrls.items():
        #             val = self.samples[_id].meta[key]
        #             if not val == item:
        #                 msg = '{}: control sample {} does not have {}:{}. \
        #                     instead, has {}'.format(self.id, _id, key, item, val)
        #                 raise Exception(msg)
        #     else:
        #         # check if non-controls have different factors
        #         for key, item in self.ctrl_ftrs.items():
        #             val = self.samples[_id].meta[key]
        #             if val == item:
        #                 msg = '{}: non-control sample {} has {}:{}'.format(self.id, _id, key, item)
        #                 raise Exception(msg)



    def write_matrix(self):
        """
        Writes rna_seq_info.txt
        """
        # First, let's construct a DataFrame with just one column.
        sample_names = []
        for sample_id in self.samples:
            sample_names.append(self.samples[sample_id].name.replace(' ', '_'))
        df = pd.DataFrame(sample_names, columns=['sample'])
        df.set_index('sample', inplace=True)

        # Add a column for each factor.
        for i in range(self.design):
            column = []
            control = self.controls[i]
            control_name = control['name']
            control_value = control['value']

            for sample_id, sample in self.samples.items():
                name = sample.name
                factor_value = sample.meta['chars'][control_name]

                # append a_ if this is a control. otherwise append b_
                if (factor_value == control_value):
                    factor_value = 'a_' + factor_value
                else:
                    factor_value = 'b_' + factor_value

                column.append([name, factor_value.replace(' ', '_')])

            col = pd.DataFrame(column, columns=['sample', control_name.replace(' ', '_')])
            col.set_index('sample', inplace=True)
            df = pd.concat([df, col], axis=1, sort=True)

        df.to_csv('{}/rna_seq_info.txt'.format(self.diff_dir), sep=' ', index=True)\

    def write_soft(self):
        """
        Writes project in SOFT format for GEO submission.
        """
        def format_indicator(indicator, value):
            return '^{} = {}\n'.format(indicator, value)

        def format_attribute(attribute, value):
            return '!{} = {}\n'.format(attribute, value)

        def write_series(f):
            """
            Writes SERIES in SOFT format.
            """
            f.write(format_indicator('SERIES', self.id))
            f.write(format_attribute('Series_title', self.meta['title']))
            f.write(format_attribute('Series_summary', self.meta['abstract']))
            f.write(format_attribute('Series_overall_design', ''))

            for cont in self.meta['contributors']:
                f.write(format_attribute('Series_contributor', cont))

            f.write(format_attribute('Series_supplementary_file', ''))

            for sample_id in self.samples:
                f.write(format_attribute('Series_sample_id', sample_id))

            supplementary = []
            diff_path = '{}/{}'.format(self.diff_dir, sample.name)
            diff_files = os.listdir(diff_path)
            for diff_file in diff_files:
                if not diff_file.endswith(('out.txt', '.rds', '.R')):
                    supplementary.append('{}/{}'.format(diff_path, diff_file))

            for file in supplementary:
                f.write(format_attribute('Series_supplementary_file', file))


        def write_sample(f, sample):
            """
            Writes the given sample in SOFT format.
            """
            f.write(format_indicator('SAMPLE', sample.id))
            f.write(format_attribute('Sample_type', 'SRA'))
            f.write(format_attribute('Sample_title', sample.meta['title']))
            f.write(format_attribute('Sample_source_name', sample.meta['chars']['tissue']))
            f.write(format_attribute('Sample_organism', sample.organism.replace('_', ' ').capitalize()))

            to_exclude = [
                'growth conditions',
                'library preparation',
                'sequenced molecules',
                'miscellaneous'
            ]

            for char, value in sample.meta['chars'].items():
                if char not in to_exclude:
                    f.write(format_attrubute('Sample_characteristics', '{}: {}'.format(char, value)))

            f.write(format_attribute('Sample_molecule', sample.meta['chars']['sequenced molecules']))
            f.write(format_attribute('Sample_growth_protocol', sample.meta['chars']['growth conditions']))
            f.write(format_attribute('Sample_library_construction_protocol', sample.meta['chars']['library preparation']))
            f.write(format_attribute('Sample_library_strategy', 'RNA-Seq'))
            f.write(format_attribute('Sample_data_processing', ''))
            f.write(format_attribute('Sample_description', sample.meta['description']))

            # Write raw files.
            if sample.type == 1:
                # construct values
                files = []
                types = []
                md5s = []
                lengths = []

                for path, read in sample.reads:
                    fname = os.path.basename(path)
                    ext = os.path.splitext(fname)[1]
                    files.append(path)
                    types.append(ext)
                    md5s.append(read['md5'])
                    lengths.append(sample.length)

                f.write(format_attribute('Sample_raw_file_name_run1', ', '.join(files)))
                f.write(format_attribute('Sample_raw_file_type_run1', ', '.join(types)))
                f.write(format_attribute('Sample_raw_file_checksum_run1', ', '.join(md5s)))
                f.write(format_attribute('Sample_raw_file_read_length_run1', ', '.join(lengths)))
                f.write(format_attribute('Sample_raw_file_standard_deviation_run1', sample.stdev))
                f.write(format_attribute('Sample_raw_file_instrument_model_run1', ''))
            elif sample.type == 2:
                pass

            # write processed data files.
            processed = []
            quant_path = '{}/{}'.format(self.align_dir, sample.name)

            quant_files = os.listdir(quant_path)
            for quant_file in quant_files:
                processed.append('{}/{}'.format(quant_path, quant_file))

            for processed_file in processed:
                f.write(format_attribute('Sample_processed_data_file', processed_file))

        with open('{}/seq_info.txt'.format(self.dir), 'w') as f:
            write_series(f)

            for sample_id, sample in self.samples:
                write_sample(f, sample)

    def write_geo_submission_form(self):
        """
        Writes the geo submission form for this project.
        """
        def find_header(sheet, header):
            """
            Finds row number of the given header.
            """
            for row in sheet.iter_rows(min_row=1, max_col=1):
                for cell in row:
                    if cell.value == header:
                        return cell.row

        def get_series_df():
            """
            Get a pandas dataframe for the SERIES category.
            """
            rows = []
            rows.append(['title', self.meta['title']])
            rows.append(['summary', self.meta['abstract']])
            rows.append(['overall design', ''])

            for contributor in self.meta['contributors']:
                rows.append(['contributor', contributor])

            rows.append(['supplementary file', ''])
            rows.append(['SRA_center_name_code', self.meta['SRA_center_code']])

            return pd.DataFrame(rows, columns=['SERIES', 'values']).set_index('SERIES')

        def get_samples_df():
            """
            Gets a pandas dataframe for the SAMPLES category.
            """
            rows = []

            # Extract all characteristics and check if PROTOCOLS are the same
            # across all samples. If they aren't they must be included
            # as additional columns for each sample.
            excl_from_chars = [
                'growth conditions',
                'library preparation',
                'sequenced molecules',
                'miscellaneous'
            ]
            chars = []
            for sample_id, sample in self.samples.items():
                for char in sample.meta['chars']:
                    if char not in chars:
                        chars.append(char)

            headers = [
                'Sample name',
                'title',
                'source name',
                'organism'
            ]
            for char in chars:
                headers.append('characteristics: ' + char)
            headers += [
                'molecule',
                'description'
            ]

            # For each header, construct a row.
            ids = self.samples.keys()
            ids = sorted(ids)
            columns = ['SAMPLE'] + ids
            rows = []
            for header in headers:
                row = []
                row.append(header)

                for sample_id in ids:
                    sample = self.samples[id]

                    if header == 'Sample name':
                        row.append(sample.id)
                    elif header == 'title':
                        row.append(sample.name)
                    elif header == 'source name':
                        row.append(sample.meta['chars']['tissue'])
                    elif header == 'organism':
                        row.append(sample.organism.replace('_', ' ').capitalize())
                    elif header.startswith('characteristics:'):
                        char = header.split(': ')[1]
                        row.append(sample.meta['chars'][char])
                    elif header == 'molecule':
                        row.append(sample.meta['chars']['sequenced molecules'])
                    elif header == 'description':
                        row.append(sample.meta['description'])

                rows.append(row)

            # make df
            df = pd.DataFrame(rows, columns=columns)

                #



        def write_series(sheet):
            """
            Write data into the SERIES category.
            """
            header = 'SERIES'
            start = find_header(sheet, header)

            # Then, iterate
            for row in sheet.iter_rows(min_row=start, max_col=2):
                cells = list(row)
                cell_1 = cells[0]
                cell_2 = cells[1]

                # current row number and value
                row_n = cell_1.row
                val = cell_1.value

                cont_done = False
                if val == 'title':
                    cell_2.value = self.meta['title']
                elif val == 'summary':
                    cell_2.value = self.meta['abstract']
                elif val == 'overall design':
                    cell_2.value = ''
                elif val == 'contributor' and not cont_done:
                    # First, deal with first contributor.
                    first = self.meta['contributors'][0]
                    cell_2.value = first

                    # Then, deal with the rest of the contributors.
                    if len(self.meta['contributors'] > 1):
                        for cont in self.meta['contributors'][1:]:
                            row_n += 1

                            # First, make a copy of this row.
                            contributor_row = [copy(cell_1), copy(cell_2)]

                            # insert a row after this row.
                            sheet.insert_rows(row_n, amount=1)

                            # Then, add the cells.
                            contributor_row[0].row = row_n
                            contributor_row[1].row = row_n
                            contributor_row[1].value = cont
                            sheet._add_cell(contributor_row[0])
                            sheet._add_cell(contributor_row[1])

                    cont_done = True

                elif val == 'supplementary file':
                    continue

                elif val == 'SRA_center_name_code':
                    if self.meta['SRA_center_code']:
                        cell_2.value = self.meta['SRA_center_code']

        def write_samples(sheet):
            """
            Write data into the SAMPLES category.
            """
            header = 'SAMPLES'
            start = find_header(sheet, header)

            for row in sheet.iter_rows(min_row=start, max_col=1):
                cells = list(row)
                cell = cells[0]
                val = cell.value

                if val == 'Sample name':
                    row_n = cell.row

        def write_protocols(sheet):
            """
            Write data into the PROTOCOLS category.
            """
            pass

        def write_pipeline(sheet):
            """
            Write data into the DATA PROCESSING PIPELINE category.
            """
            pass

        def write_processed(sheet):
            """
            Write data into the PROCESSED DATA FILES category.
            """
            pass

        def write_raw(sheet):
            """
            Write data into the RAW FILES and PAIRED-END EXPERIMENTS categories.
            """
            pass

        # Load template.
        wb = opx.load_workbook(Alaska.GEO_FILE)
        sheet = wb['METADATA TEMPLATE']

        # Populate categories.
        write_series(sheet)
        write_samples(sheet)
        write_protocols(sheet)
        write_pipeline(sheet)
        write_processed(sheet)
        write_raw(sheet)

        # Save back.
        wb.save(Alaska.GEO_FILE)

    def save(self, folder=None):
        """
        Save project to JSON.
        """
        if folder is None: # if folder not given, save to project root
            path = self.dir
        else:
            path = '{}/{}'.format(self.dir, folder)

        with open('{}/{}.json'.format(path, self.id), 'w') as f:
            json.dump(self.__dict__, f, default=self.encode_json, indent=4)


    def load(self, folder=None):
        """
        Loads project from JSON.
        """
        if folder is None: # if folder not given, load from project root
            path = self.dir
        else:
            path = '{}/{}'.format(self.dir, folder)

        with open('{}/{}.json'.format(path, self.id), 'r') as f:
            loaded = json.load(f)

        if not loaded['id'] == self.id:
            raise Exception('{}: JSON id {} does not match object id'
                        .format(self.id, loaded['id']))

        for key, item in loaded.items():
            if key == 'samples':
                # AlaskaSample object must be created for samples
                samples = {}
                for _id, vals in item.items():
                    sample = AlaskaSample(_id)
                    sample.__dict__ = vals
                    samples[_id] = sample
                # set samples
                self.samples = samples
            else:
                setattr(self, key, item)
